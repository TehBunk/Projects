from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

WORKBOOK = None
MONEY = None

# work_book.sheetnames # GET ALL SHEETS
# work_book.active;  # GET ACTIVE SHEET
# work_book.title # GET TITLE OF SHEET
# work_book.title = "TEST" # SET TITLE OF SHEET
# work_book.['Sheet1']; # GET SHEET BY NAME
# work_book.create_sheet("test") # CREATE NEW SHEET


# work_sheet['A1'].value # GET VALUE OF CELL
# work_sheet['A1'].value = 'Name' # CHANGE VALUE OF CELL
# work_sheet.append(["test", "test"]) # ADD LOT OF DATA
# work_sheet.merge_cells("A1:D1")
# work_sheet.unmerge_cells("A1:D1")
# work_sheet.get_column_letter(0) will return A
# work_sheet.insert_rows(8)
# work_sheet.delete_rows(8)

def create_stock_sheet(ticker, income_data, balance_data, cashflow_data):
    sheet = create_sheet(ticker.upper())

    for row in range(len(income_data)):
        for col in range(len(income_data[0])):
            data = income_data[row][col]
            if (col == 0): # desc
                sheet.cell(row + 1, col + 1, data)
            else: # currency
                if (type(data) is str and "$" in data):  
                    sheet.cell(row + 1, col + 2, data)
                    sheet.cell(row + 1, col + 2).number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'
                else:  # number
                    sheet.cell(row + 1, col + 2, float(data))


def create_sheet(name):
    return WORKBOOK.create_sheet(name)


def load_excel():
    global WORKBOOK
    global MONEY
    WORKBOOK = load_workbook('Pillars-Scraper\Book1.xlsx')


def save_excel():
    WORKBOOK.save('Pillars-Scraper\Book1.xlsx')
